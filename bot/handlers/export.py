# bot/export.py
import os
import tempfile
import datetime as dt
from dataclasses import dataclass
from typing import List, Optional

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font

from ..utils.timezone import format_moscow, to_moscow

@dataclass
class AnswerRow:
    number: int
    question: str
    qtype: str
    answer: str
    comment: Optional[str] = None
    score: Optional[float] = None
    photo_path: Optional[str] = None

@dataclass
class AttemptData:
    attempt_id: int
    checklist_name: str
    user_name: str
    company_name: Optional[str]
    department: Optional[str] = None
    submitted_at: dt.datetime
    answers: List[AnswerRow]
    total_score: Optional[float] = None
    total_max: Optional[float] = None
    percent: Optional[float] = None
    is_scored: bool = False

def _register_font():
    font_path = os.path.join("assets", "fonts", "DejaVuSans.ttf")
    if os.path.exists(font_path):
        try:
            pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
            return "DejaVuSans"
        except Exception:
            pass
    return None

def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def _fmt_dt(d: dt.datetime) -> str:
    return format_moscow(d, "%Y-%m-%d %H:%M")

def export_attempt_to_pdf(filename: str, data: AttemptData):
    font_name = _register_font()
    styles = getSampleStyleSheet()
    if font_name:
        for k in styles.byName:
            styles.byName[k].fontName = font_name
        title_style = ParagraphStyle(name="TitleCustom", parent=styles["Title"], fontName=font_name)
        h2_style = ParagraphStyle(name="H2Custom", parent=styles["Heading2"], fontName=font_name)
        normal_style = ParagraphStyle(name="NormalCustom", parent=styles["Normal"], fontName=font_name)
    else:
        title_style = styles["Title"]
        h2_style = styles["Heading2"]
        normal_style = styles["Normal"]

    doc = SimpleDocTemplate(filename, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    elements = []

    elements.append(Paragraph("Отчёт по чек-листу", title_style))
    elements.append(Spacer(1, 12))
    elements.append(Paragraph(f"<b>Чек-лист:</b> {data.checklist_name}", normal_style))
    elements.append(Paragraph(f"<b>Сотрудник:</b> {data.user_name}", normal_style))
    if data.company_name:
        elements.append(Paragraph(f"<b>Компания:</b> {data.company_name}", normal_style))
    elements.append(Paragraph(f"<b>Дата прохождения:</b> {_fmt_dt(data.submitted_at)}", normal_style))
    elements.append(Spacer(1, 16))

    table_data = [["№", "Вопрос", "Тип", "Ответ", "Комментарий", "Балл", "Фото"]]
    for row in data.answers:
        photo_label = os.path.basename(row.photo_path) if row.photo_path else ""
        table_data.append([
            row.number,
            row.question,
            row.qtype,
            row.answer,
            row.comment or "",
            "" if row.score is None else row.score,
            photo_label
        ])

    table = Table(table_data, colWidths=[28, 210, 60, 120, 120, 40, 70])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#333333")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('FONTNAME', (0,0), (-1,0), font_name or 'Helvetica-Bold'),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('GRID', (0,0), (-1,-1), 0.4, colors.grey),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTSIZE', (0,0), (-1,-1), 9),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.Color(0.98,0.98,0.98)]),
    ]))
    elements.append(table)
    elements.append(Spacer(1, 14))

    images = [r.photo_path for r in data.answers if r.photo_path and os.path.exists(r.photo_path)]
    if images:
        elements.append(Paragraph("Фотоприложения:", h2_style))
        elements.append(Spacer(1, 8))
        for p in images[:8]:
            try:
                elements.append(RLImage(p, width=200, height=120))
                elements.append(Spacer(1, 6))
                elements.append(Paragraph(os.path.basename(p), normal_style))
                elements.append(Spacer(1, 8))
            except Exception:
                continue

    doc.build(elements)

def export_attempt_to_excel(filename: str, data: AttemptData):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ответы"

    ws.append(["№", "Вопрос", "Тип", "Ответ", "Комментарий", "Балл", "Фото"])
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    for row in data.answers:
        ws.append([
            row.number,
            row.question,
            row.qtype,
            row.answer,
            row.comment or "",
            row.score if row.score is not None else "",
            os.path.basename(row.photo_path) if row.photo_path else ""
        ])

    _auto_fit_columns(ws)

    ws2 = wb.create_sheet("Сводка")
    ws2.append(["Параметр", "Значение"])
    ws2["A1"].font = header_font
    ws2["B1"].font = header_font

    ws2.append(["Чек-лист", data.checklist_name])
    ws2.append(["Сотрудник", data.user_name])
    if data.company_name:
        ws2.append(["Компания", data.company_name])
    ws2.append(["Дата прохождения", _fmt_dt(data.submitted_at)])
    ws2.append(["Всего вопросов", len(data.answers)])
    total_score = sum([r.score for r in data.answers if isinstance(r.score, (int, float))])
    ws2.append(["Суммарный балл", total_score])

    _auto_fit_columns(ws2)
    wb.save(filename)

def export_attempt_to_files(tmp_dir: Optional[str], data: AttemptData):
    base_dir = tmp_dir or tempfile.gettempdir()
    safe_user = data.user_name.replace(" ", "_")
    safe_check = data.checklist_name.replace(" ", "_")
    local_dt = to_moscow(data.submitted_at) or data.submitted_at
    stamp = local_dt.strftime("%Y%m%d_%H%M")
    pdf_path = os.path.join(base_dir, f"report_{safe_check}_{safe_user}_{stamp}.pdf")
    xlsx_path = os.path.join(base_dir, f"report_{safe_check}_{safe_user}_{stamp}.xlsx")

    export_attempt_to_pdf(pdf_path, data)
    export_attempt_to_excel(xlsx_path, data)
    return pdf_path, xlsx_path
