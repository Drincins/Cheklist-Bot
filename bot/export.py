# bot/export.py
import logging
import os
import tempfile
import datetime as dt
import math
from typing import Optional, List
from xml.sax.saxutils import escape

from dotenv import load_dotenv
load_dotenv()

from reportlab.lib.pagesizes import A4
from reportlab.platypus import (
    SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage
)
from reportlab.platypus.flowables import Flowable
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage

from .report_data import AnswerRow, AttemptData, SectionResult
from .utils.timezone import to_moscow, format_moscow

logger = logging.getLogger(__name__)

# === Утилиты ===
def _register_font():
    font_path = os.getenv("PDF_FONT_PATH")
    if font_path:
        if os.path.exists(font_path):
            try:
                pdfmetrics.registerFont(TTFont("DejaVuSans", font_path))
                logger.info("[PDF] Using font from PDF_FONT_PATH: %s", font_path)
                return "DejaVuSans"
            except Exception as e:
                logger.warning("[PDF] Failed to register font from PDF_FONT_PATH: %s", e)
        else:
            logger.warning("[PDF] PDF_FONT_PATH points to missing file: %s", font_path)

    candidates = [
        "assets/fonts/DejaVuSans.ttf",
        "/Library/Fonts/DejaVuSans.ttf",
        "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf",
        "C:\\Windows\\Fonts\\DejaVuSans.ttf",
    ]
    for p in candidates:
        if os.path.exists(p):
            try:
                pdfmetrics.registerFont(TTFont("DejaVuSans", p))
                logger.info("[PDF] Using fallback font: %s", p)
                return "DejaVuSans"
            except Exception as e:
                logger.warning("[PDF] Failed to register fallback font %s: %s", p, e)

    logger.warning("[PDF] No font registered, fallback to Helvetica (may break Cyrillic)")
    return None

def _auto_fit_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

def _fmt_dt(d: dt.datetime) -> str:
    return format_moscow(d, "%Y-%m-%d %H:%M")


def _fmt_number(value: Optional[float]) -> str:
    if value is None:
        return ""
    return ("{:.2f}".format(value)).rstrip("0").rstrip(".")


# === PDF ===
def export_attempt_to_pdf(filename: str, data: AttemptData):
    font_name = _register_font()
    styles = getSampleStyleSheet()
    if font_name:
        for k in styles.byName:
            styles.byName[k].fontName = font_name

    title_style = ParagraphStyle(
        name="TitleCustom",
        parent=styles["Title"],
        fontSize=20,
        leading=24,
        textColor=colors.HexColor("#1F3B4D"),
        alignment=1,
    )
    h2_style = ParagraphStyle(
        name="H2Custom",
        parent=styles["Heading2"],
        textColor=colors.HexColor("#1F3B4D"),
    )
    normal_style = ParagraphStyle(name="NormalCustom", parent=styles["Normal"], leading=14)
    label_style = ParagraphStyle(name="Label", parent=normal_style, textColor=colors.HexColor("#1F3B4D"), fontSize=11)
    header_cell_style = ParagraphStyle(
        name="HeaderCell",
        parent=normal_style,
        textColor=colors.whitesmoke,
        fontName=font_name or 'Helvetica-Bold',
    )
    result_style = ParagraphStyle(
        name="Result",
        parent=normal_style,
        fontSize=16,
        leading=20,
        textColor=colors.HexColor("#0B6623"),
        alignment=1,
        spaceBefore=12,
        spaceAfter=18,
    )
    section_title_style = ParagraphStyle(
        name="SectionTitle",
        parent=styles["Heading3"],
        fontName=font_name or styles["Heading3"].fontName,
        fontSize=14,
        leading=18,
        textColor=colors.HexColor("#1F3B4D"),
        spaceBefore=12,
        spaceAfter=6,
    )
    section_score_style = ParagraphStyle(
        name="SectionScore",
        parent=normal_style,
        fontSize=10,
        leading=12,
        textColor=colors.HexColor("#4A4A4A"),
        spaceBefore=4,
        spaceAfter=6,
    )

    doc = SimpleDocTemplate(filename, pagesize=A4, leftMargin=24, rightMargin=24, topMargin=24, bottomMargin=24)
    elements = []

    logo_path_env = os.getenv("PDF_LOGO_PATH")
    logo_candidates = [logo_path_env] if logo_path_env else []
    logo_candidates.append(os.path.join("assets", "logo.png"))
    logo_flowable = None
    for candidate in logo_candidates:
        if candidate and os.path.exists(candidate):
            try:
                logo = RLImage(candidate)
                max_w, max_h = 130, 46
                ratio = min(max_w / logo.drawWidth, max_h / logo.drawHeight, 1.0)
                logo.drawWidth *= ratio
                logo.drawHeight *= ratio
                logo_flowable = logo
                break
            except Exception as e:
                logger.warning("[PDF] logo load failed for %s: %s", candidate, e)
                continue

    title_box: Flowable
    if logo_flowable:
        title_box = Table(
            [[logo_flowable, Paragraph("<b>Отчёт по чек-листу</b>", title_style)]],
            colWidths=[logo_flowable.drawWidth, doc.width - logo_flowable.drawWidth],
            hAlign='LEFT'
        )
        title_box.setStyle(TableStyle([
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
            ('ALIGN', (0,0), (0,0), 'LEFT'),
            ('ALIGN', (1,0), (1,0), 'LEFT'),
            ('LEFTPADDING', (0,0), (-1,-1), 0),
            ('RIGHTPADDING', (0,0), (-1,-1), 0),
            ('TOPPADDING', (0,0), (-1,-1), 0),
            ('BOTTOMPADDING', (0,0), (-1,-1), 0),
        ]))
        elements.append(title_box)
    else:
        elements.append(Paragraph("<b>Отчёт по чек-листу</b>", title_style))
    elements.append(Spacer(1, 14))

    info_rows = [
        [Paragraph("<b>Чек-лист</b>", label_style), Paragraph(data.checklist_name, normal_style)],
        [Paragraph("<b>Сотрудник</b>", label_style), Paragraph(data.user_name, normal_style)],
    ]
    if data.company_name:
        info_rows.append([Paragraph("<b>Компания</b>", label_style), Paragraph(data.company_name, normal_style)])
    if data.department:
        info_rows.append([Paragraph("<b>Подразделение</b>", label_style), Paragraph(data.department, normal_style)])
    info_rows.append([Paragraph("<b>Дата прохождения</b>", label_style), Paragraph(_fmt_dt(data.submitted_at), normal_style)])

    info_table = Table(info_rows, colWidths=[120, doc.width - 120])
    info_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (0,-1), colors.HexColor("#E8F1F5")),
        ('BOX', (0,0), (-1,-1), 0.6, colors.HexColor("#C6D7E2")),
        ('INNERGRID', (0,0), (-1,-1), 0.4, colors.HexColor("#C6D7E2")),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('LEFTPADDING', (0,0), (-1,-1), 6),
        ('RIGHTPADDING',(0,0), (-1,-1), 6),
        ('TOPPADDING',  (0,0), (-1,-1), 4),
        ('BOTTOMPADDING',(0,0), (-1,-1), 4),
    ]))
    elements.append(info_table)
    elements.append(Spacer(1, 14))

    # нормализация ответа только для yes/no
    def _norm_answer_for_row(row: AnswerRow) -> str:
        val = row.answer
        if val is None or str(val).strip() == "":
            return "*пусто*"
        qtype = (row.qtype or "").lower().strip()
        if qtype in {"yesno", "boolean", "bool", "yn"}:
            s = str(val).strip().lower()
            if s in {"yes", "да", "true"}:  return "Да"
            if s in {"no", "нет", "false"}: return "Нет"
            if s == "1": return "Да"
            if s == "0": return "Нет"
        return str(val)
    
    def _fmt_score(row) -> str:
        wt = getattr(row, "weight", None)
        sc = getattr(row, "score", None)
        if wt is None:
            return ""
        try:
            wt = float(wt)
        except Exception:
            return ""
        if sc is None:
            sc = 0.0
        else:
            sc = float(sc)

        def f(x: float) -> str:
            return str(int(x)) if abs(x - int(x)) < 1e-9 else f"{x:.2f}".rstrip("0").rstrip(".")
        return f"{f(sc)}/{f(wt)}"


    # помощник фото в ячейке
    def _image_cell(path: str, max_w: int = 45, max_h: int = 45):
        try:
            if path and os.path.exists(path):
                img = RLImage(path)
                ratio = min(max_w / img.drawWidth, max_h / img.drawHeight, 1.0)
                img.drawWidth *= ratio
                img.drawHeight *= ratio
                return img
        except Exception as e:
            logger.warning("[PDF] inline image failed for %s: %s", path, e)
        return ""

    def P(text: str):
        return Paragraph((text or "").replace("\n", "<br/>"), normal_style)

    def PH(text: str):
        return Paragraph((text or "").replace("\n", "<br/>"), header_cell_style)

    # колонки: № | Вопрос | Ответ | Комментарий | Балл | Фото  (БЕЗ «Тип»)
    headers = ["№", "Вопрос", "Ответ", "Комментарий", "Балл/Вес", "Фото"]

    questions_plain = [headers[1]] + [row.question or "" for row in data.answers]
    answers_plain = [headers[2]] + [_norm_answer_for_row(row) for row in data.answers]
    comments_plain = [headers[3]] + [row.comment or "" for row in data.answers]
    scores_plain = [headers[4]] + [_fmt_score(row) for row in data.answers]
    numbers_plain = [headers[0]] + [str(row.number) for row in data.answers]

    photo_draw_widths: list[float] = [0.0]
    for row in data.answers:
        tmp_photo = _image_cell(row.photo_path) if row.photo_path else ""
        if isinstance(tmp_photo, RLImage):
            photo_draw_widths.append(tmp_photo.drawWidth)
        else:
            photo_draw_widths.append(0.0)

    from reportlab.pdfbase.pdfmetrics import stringWidth

    def _measure(strings, min_width, max_width, padding=12) -> float:
        nonlocal font_name
        font = font_name or 'Helvetica'
        widths = []
        for s in strings:
            text = str(s or "")
            text = text.replace('\n', ' ')
            widths.append(stringWidth(text, font, 9))
        raw = (max(widths) if widths else min_width) + padding
        return max(min_width, min(raw, max_width))

    # базовые ограничения (pt)
    num_width = _measure(numbers_plain, 26, 40)
    score_width = _measure(scores_plain, 45, 70)
    photo_width = max(max(photo_draw_widths), 45)
    photo_width = max(45, min(photo_width + 8, 80))

    available = doc.width - (num_width + score_width + photo_width)

    min_text_width = 150
    if available < min_text_width:
        overflow = min_text_width - available
        photo_width = max(45, photo_width - overflow)
        available = doc.width - (num_width + score_width + photo_width)
        if available < min_text_width:
            available = min_text_width

    q_width_raw = _measure(questions_plain, 120, 320, padding=18)
    a_width_raw = _measure(answers_plain, 90, 200, padding=18)
    c_width_raw = _measure(comments_plain, 110, 260, padding=18)

    text_raw_total = q_width_raw + a_width_raw + c_width_raw

    if text_raw_total > available:
        factor = available / text_raw_total if text_raw_total else 1
        q_width = max(120, q_width_raw * factor)
        a_width = max(90, a_width_raw * factor)
        c_width = max(110, c_width_raw * factor)
    else:
        leftover = available - text_raw_total
        weights = [max(q_width_raw - 120, 1), max(a_width_raw - 90, 1), max(c_width_raw - 110, 1)]
        weight_sum = sum(weights)
        if weight_sum == 0:
            increment = leftover / 3
            q_width = q_width_raw + increment
            a_width = a_width_raw + increment
            c_width = c_width_raw + increment
        else:
            q_width = q_width_raw + leftover * (weights[0] / weight_sum)
            a_width = a_width_raw + leftover * (weights[1] / weight_sum)
            c_width = c_width_raw + leftover * (weights[2] / weight_sum)

    col_widths = [num_width, q_width, a_width, c_width, score_width, photo_width]

    total_width = sum(col_widths)
    if total_width > doc.width:
        factor = doc.width / total_width
        col_widths = [w * factor for w in col_widths]

    table_style = TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor("#1F3B4D")),
        ('TEXTCOLOR', (0,0), (-1,0), colors.whitesmoke),
        ('ALIGN', (0,0), (-1,0), 'CENTER'),
        ('FONTNAME', (0,0), (-1,0), font_name or 'Helvetica-Bold'),

        ('FONTNAME', (0,1), (-1,-1), font_name or 'Helvetica'),
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ('FONTSIZE', (0,0), (-1,-1), 9),

        ('LEFTPADDING',  (0,0), (-1,-1), 6),
        ('RIGHTPADDING', (0,0), (-1,-1), 6),
        ('TOPPADDING',   (0,0), (-1,-1), 3),
        ('BOTTOMPADDING',(0,0), (-1,-1), 3),

        ('GRID', (0,0), (-1,-1), 0.4, colors.HexColor("#C7CFD9")),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.whitesmoke, colors.Color(0.96,0.97,0.99)]),

        ('ALIGN', (0,1), (0,-1), 'CENTER'),
        ('ALIGN', (4,1), (4,-1), 'CENTER'),
        ('ALIGN', (5,1), (5,-1), 'CENTER'),
    ])

    sections_to_render: List[SectionResult]
    if data.sections:
        sections_to_render = [sec for sec in data.sections if sec.answers]
    else:
        sections_to_render = []

    use_sections = len(sections_to_render) > 1
    if not sections_to_render:
        sections_to_render = [SectionResult(title=None, answers=data.answers)]

    for idx, section in enumerate(sections_to_render, start=1):
        answers_in_section = section.answers
        if not answers_in_section:
            continue

        if use_sections:
            title_text = section.title or "Без раздела"
            elements.append(Paragraph(
                f"📂 Блок {idx}: <b>{escape(title_text)}</b>",
                section_title_style,
            ))

        section_table_rows = [[
            PH(headers[0]),
            PH(headers[1]),
            PH(headers[2]),
            PH(headers[3]),
            PH(headers[4]),
            PH(headers[5]),
        ]]

        for row in answers_in_section:
            section_table_rows.append([
                row.number,
                P(row.question),
                P(_norm_answer_for_row(row)),
                P(row.comment or ""),
                _fmt_score(row),
                _image_cell(row.photo_path) if row.photo_path else "",
            ])

        section_table = Table(
            section_table_rows,
            colWidths=col_widths,
            repeatRows=1,
            hAlign='LEFT'
        )
        section_table.setStyle(table_style)
        elements.append(section_table)

        if use_sections and data.is_scored and section.total_score is not None and section.total_max is not None:
            percent_text = f" ({_fmt_number(section.percent)}%)" if section.percent is not None else ""
            section_summary = (
                f"Результат блока: {_fmt_number(section.total_score)} из "
                f"{_fmt_number(section.total_max)} баллов{percent_text}"
            )
            elements.append(Paragraph(section_summary, section_score_style))

        elements.append(Spacer(1, 12))

    if data.is_scored and data.total_score is not None and data.total_max is not None:
        percent_text = f" ({_fmt_number(data.percent)}%)" if data.percent is not None else ""
        result_text = (
            f"Набрано {_fmt_number(data.total_score)} из "
            f"{_fmt_number(data.total_max)} баллов{percent_text}"
        )
        elements.append(Paragraph(result_text, result_style))
        elements.append(Spacer(1, 12))

    # (опционально) крупные фотоприложения ниже
    images = [(r.photo_path, r.photo_label or f"Вопрос №{r.number}") for r in data.answers if r.photo_path and os.path.exists(r.photo_path)]
    if images:
        elements.append(Paragraph("Фотоприложения:", h2_style))
        elements.append(Spacer(1, 8))
        max_w, max_h = 380, 240
        for p, label in images[:8]:
            try:
                img = RLImage(p)
                ratio = min(max_w / img.drawWidth, max_h / img.drawHeight, 1.0)
                img.drawWidth *= ratio
                img.drawHeight *= ratio
                elements.append(img)
                elements.append(Spacer(1, 6))
                elements.append(Paragraph(label, normal_style))
                elements.append(Spacer(1, 8))
            except Exception as e:
                logger.warning("[PDF] image add failed for %s: %s", p, e)
                continue

    doc.build(elements)


# === Excel ===
def export_attempt_to_excel(filename: str, data: AttemptData):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ответы"

    # Шапка (БЕЗ «Тип»)
    ws.append(["№", "Вопрос", "Ответ", "Комментарий", "Балл/Вес", "Фото"])
    header_font = Font(bold=True)
    for cell in ws[1]:
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Базовые ширины колонок
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 60
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 40
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 18  # фото

    # Нормализация ответа только для yes/no
    def _norm_answer_for_row(row: AnswerRow) -> str:
        val = row.answer
        if val is None or str(val).strip() == "":
            return "*пусто*"
        qtype = (row.qtype or "").lower().strip()
        if qtype in {"yesno", "boolean", "bool", "yn"}:
            s = str(val).strip().lower()
            if s in {"yes", "да", "true"}:  return "Да"
            if s in {"no", "нет", "false"}: return "Нет"
            if s == "1": return "Да"
            if s == "0": return "Нет"
        return str(val)

    def _fmt_score(row) -> str:
        if row.weight is None:
            return ""
        sc = 0.0 if row.score is None else float(row.score)
        wt = float(row.weight)
        def f(x):
            return str(int(x)) if abs(x - int(x)) < 1e-9 else f"{x:.2f}".rstrip('0').rstrip('.')
        return f"{f(sc)}/{f(wt)}"


    # Настройки миниатюр
    max_w_px = 160
    max_h_px = 120
    excel_px_to_pts = 0.75  # 1 px ≈ 0.75 pt

    def _make_thumb(src_path: str) -> tuple[Optional[str], int, int]:
        try:
            if not (src_path and os.path.exists(src_path)):
                return None, 0, 0
            im = PILImage.open(src_path)
            im = im.convert("RGB")
            im.thumbnail((max_w_px, max_h_px))
            tmp = os.path.join(tempfile.gettempdir(), f"xlthumb_{os.path.basename(src_path)}")
            im.save(tmp, format="JPEG", quality=85)
            return tmp, im.width, im.height
        except Exception as e:
            logger.warning("[XLSX] thumb error for %s: %s", src_path, e)
            return None, 0, 0

    def _estimate_row_height(question: str, answer: str, comment: str) -> float:
        def lines(text: str, width_chars: int) -> int:
            if not text:
                return 1
            count = 0
            for part in text.splitlines() or [""]:
                clean = part.strip() or ""
                length = len(clean)
                count += max(1, math.ceil(length / max(width_chars, 1)))
            return count

        q_lines = lines(question, 60)
        a_lines = lines(answer, 30)
        c_lines = lines(comment, 40)
        base_lines = max(q_lines, a_lines, c_lines)
        return 15 * base_lines + 6

    sections_to_render: List[SectionResult]
    if data.sections:
        sections_to_render = [sec for sec in data.sections if sec.answers]
    else:
        sections_to_render = []

    use_sections = len(sections_to_render) > 1
    if not sections_to_render:
        sections_to_render = [SectionResult(title=None, answers=data.answers)]

    tmp_thumbs: list[str] = []
    photo_refs: list[tuple[str, str]] = []
    try:
        current_row = 2

        for sec_idx, section in enumerate(sections_to_render, start=1):
            answers_in_section = section.answers
            if not answers_in_section:
                continue

            if use_sections:
                if current_row > 2:
                    current_row += 1  # пустая строка между блоками
                block_title = section.title or "Без раздела"
                ws.cell(row=current_row, column=1, value=f"Блок {sec_idx}: {block_title}")
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                ws.cell(row=current_row, column=1).font = Font(bold=True, size=12)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
                ws.cell(row=current_row, column=1).fill = PatternFill(start_color="F0F4FF", end_color="F0F4FF", fill_type="solid")
                ws.row_dimensions[current_row].height = 22
                current_row += 1

            for row in answers_in_section:
                row_idx = current_row
                norm_answer = _norm_answer_for_row(row)

                ws.cell(row=row_idx, column=1, value=row.number).alignment = Alignment(horizontal="center", vertical="top")
                ws.cell(row=row_idx, column=2, value=row.question).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=row_idx, column=3, value=norm_answer).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=row_idx, column=4, value=row.comment or "").alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(row=row_idx, column=5, value=_fmt_score(row)).alignment = Alignment(horizontal="center", vertical="top")

                approx_height = _estimate_row_height(row.question or "", norm_answer, row.comment or "")
                ws.row_dimensions[row_idx].height = max(ws.row_dimensions[row_idx].height or 0, approx_height)

                if row.photo_path and os.path.exists(row.photo_path):
                    thumb_path, w_px, h_px = _make_thumb(row.photo_path)
                    if thumb_path:
                        tmp_thumbs.append(thumb_path)
                        img = XLImage(thumb_path)
                        ws.add_image(img, f"F{row_idx}")
                        ws.row_dimensions[row_idx].height = max(
                            ws.row_dimensions[row_idx].height or 0,
                            int((h_px + 10) * excel_px_to_pts),
                        )
                    photo_refs.append((row.photo_label or f"Вопрос №{row.number}", row.photo_path))
                else:
                    ws.cell(row=row_idx, column=6, value="").alignment = Alignment(vertical="top")

                current_row += 1

            if use_sections and data.is_scored and section.total_score is not None and section.total_max is not None:
                summary_text = (
                    f"Результат блока: {_fmt_number(section.total_score)} из "
                    f"{_fmt_number(section.total_max)} баллов"
                )
                if section.percent is not None:
                    summary_text += f" ({_fmt_number(section.percent)}%)"
                ws.cell(row=current_row, column=1, value=summary_text)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=6)
                ws.cell(row=current_row, column=1).font = Font(size=10, italic=True)
                ws.cell(row=current_row, column=1).alignment = Alignment(horizontal="left", vertical="center")
                current_row += 1
    finally:
        if photo_refs:
            ws.append([])
            ws.append(["Фотоприложения", ""])
            ws[ws.max_row][0].font = header_font
            ws[ws.max_row][0].alignment = Alignment(horizontal="left", vertical="top")
            for label, path in photo_refs:
                ws.append([label, os.path.basename(path)])

        _auto_fit_columns(ws)

        # Лист "Сводка"
        ws2 = wb.create_sheet("Сводка")
        ws2.append(["Параметр", "Значение"])
        ws2["A1"].font = header_font
        ws2["B1"].font = header_font

        ws2.append(["Чек-лист", data.checklist_name])
        ws2.append(["Сотрудник", data.user_name])
        if data.company_name:
            ws2.append(["Компания", data.company_name])
        if data.department:
            ws2.append(["Подразделение", data.department])
        ws2.append(["Дата прохождения", _fmt_dt(data.submitted_at)])
        ws2.append(["Всего вопросов", len(data.answers)])
        if data.is_scored and data.total_score is not None and data.total_max is not None:
            res_text = (
                f"{_fmt_number(data.total_score)} из "
                f"{_fmt_number(data.total_max)} баллов"
            )
            if data.percent is not None:
                res_text += f" ({_fmt_number(data.percent)}%)"
            ws2.append(["Результат", res_text])
        else:
            total_score = sum([r.score for r in data.answers if isinstance(r.score, (int, float))])
            ws2.append(["Набранные баллы", _fmt_number(total_score)])

        if data.is_scored and use_sections:
            block_rows = [
                (idx, sec)
                for idx, sec in enumerate(sections_to_render, start=1)
                if sec.total_score is not None and sec.total_max is not None
            ]
            if block_rows:
                ws2.append([])
                ws2.append(["Блок", "Результат"])
                block_header_row = ws2.max_row
                ws2[f"A{block_header_row}"].font = header_font
                ws2[f"B{block_header_row}"].font = header_font
                for idx, sec in block_rows:
                    res_text = (
                        f"{_fmt_number(sec.total_score)} из "
                        f"{_fmt_number(sec.total_max)} баллов"
                    )
                    if sec.percent is not None:
                        res_text += f" ({_fmt_number(sec.percent)}%)"
                    ws2.append([f"Блок {idx}: {sec.title or 'Без раздела'}", res_text])
        _auto_fit_columns(ws2)

        wb.save(filename)

        # Чистка временных превью
        for p in tmp_thumbs:
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass


# === Универсальная обёртка (создаёт оба файла и возвращает пути) ===
def export_attempt_to_files(tmp_dir: Optional[str], data: AttemptData):
    base_dir = tmp_dir or tempfile.gettempdir()
    safe_user = data.user_name.replace(" ", "_")
    safe_check = data.checklist_name.replace(" ", "_")
    local_dt = to_moscow(data.submitted_at) or data.submitted_at
    stamp = local_dt.strftime("%Y%m%d_%H%M")
    pdf_path = os.path.join(base_dir, f"report_{safe_check}_{safe_user}_{stamp}.pdf")
    xlsx_path = os.path.join(base_dir, f"report_{safe_check}_{safe_user}_{stamp}.xlsx")

    export_attempt_to_pdf(pdf_path, data)
    export_attempt_to_excel(xlsx_path, data)
    return pdf_path, xlsx_path
